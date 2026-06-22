from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import DataProfile, Submission
from app.services.json_safety import make_json_safe
from app.services.semantic_schema import canonical_target_for_column

PROFILER_VERSION = "1.0"
MAX_SAMPLE_VALUES = 3
MAX_DISTINCT_TRACK = 25

COLUMN_ALIASES = {
    "date": "voucher_date",
    "entry no": "entry_no",
    "entry_number": "entry_no",
    "sub account": "sub_account",
    "ledger_name": "sub_account",
    "particulars": "details",
    "account class": "class",
    "account_class": "class",
    "account subclass": "sub_class",
    "account_subclass": "sub_class",
    "debit": "debit_amount",
    "credit": "credit_amount",
    "debit amount": "debit_amount",
    "credit amount": "credit_amount",
    "account code": "account_code",
}


def compute_file_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def build_data_profile_from_file(
    path: Path,
    *,
    max_preview_rows: int,
) -> tuple[dict[str, Any], list[dict[str, Any]]] | None:
    extension = path.suffix.lower()
    if extension in {".csv", ".tsv", ".xlsx", ".xls"}:
        frame, total_rows = _load_tabular_preview(path, extension, max_preview_rows=max_preview_rows)
    elif extension == ".json":
        frame, total_rows = _load_json_frame(path, max_preview_rows=max_preview_rows)
    elif extension == ".txt":
        frame, total_rows = _load_text_frame(path, max_preview_rows=max_preview_rows)
    else:
        return None

    if frame is None or frame.empty:
        return None

    frame = frame.astype(object).where(pd.notnull(frame), None)
    detected_types = _infer_detected_types(frame)
    preview_rows = _records_from_frame(frame)
    source_columns = [str(column) for column in frame.columns]
    fingerprint = compute_file_fingerprint(path)

    profile = {
        "file_fingerprint": fingerprint,
        "profiler_version": PROFILER_VERSION,
        "profile_status": "ready",
        "row_count": total_rows,
        "preview_row_count": len(preview_rows),
        "source_columns": source_columns,
        "detected_types": detected_types,
        "preview_rows": preview_rows,
        "columns": [_column_profile(frame, str(column), detected_types.get(str(column), "string")) for column in frame.columns],
    }
    safe_profile = make_json_safe(profile)
    return safe_profile, safe_profile.get("preview_rows", [])


async def load_latest_data_profile(db: AsyncSession, submission_id) -> DataProfile | None:
    stmt = (
        select(DataProfile)
        .where(DataProfile.submission_id == submission_id)
        .order_by(DataProfile.created_at.desc())
        .limit(1)
    )
    return (await db.execute(stmt)).scalars().first()


async def get_or_create_data_profile(
    db: AsyncSession,
    *,
    submission: Submission,
    path: Path,
    max_preview_rows: int,
) -> tuple[DataProfile | None, dict[str, Any], list[dict[str, Any]]]:
    built = build_data_profile_from_file(path, max_preview_rows=max_preview_rows)
    if built is None:
        return None, {}, []
    profile_payload, preview_rows = built
    fingerprint = str(profile_payload.get("file_fingerprint", "")).strip()
    profiler_version = str(profile_payload.get("profiler_version", PROFILER_VERSION)).strip() or PROFILER_VERSION

    existing_stmt = (
        select(DataProfile)
        .where(
            DataProfile.submission_id == submission.id,
            DataProfile.file_fingerprint == fingerprint,
            DataProfile.profiler_version == profiler_version,
        )
        .limit(1)
    )
    existing = (await db.execute(existing_stmt)).scalars().first()
    if existing is not None:
        return existing, make_json_safe(existing.profile_json), preview_rows

    record = DataProfile(
        submission_id=submission.id,
        file_fingerprint=fingerprint,
        profiler_version=profiler_version,
        profile_json=profile_payload,
        status=str(profile_payload.get("profile_status", "ready")),
    )
    db.add(record)
    await db.flush()
    return record, profile_payload, preview_rows


def _column_profile(frame: pd.DataFrame, column: str, detected_type: str) -> dict[str, Any]:
    series = frame[column]
    non_null = series.dropna()
    distinct_values = list(dict.fromkeys(str(value) for value in non_null.tolist() if value not in {None, ""}))
    bounded_distinct = distinct_values[:MAX_DISTINCT_TRACK]
    min_value = None
    max_value = None
    try:
        numeric_series = pd.to_numeric(non_null, errors="coerce").dropna()
        if not numeric_series.empty:
            min_value = numeric_series.min()
            max_value = numeric_series.max()
        elif not non_null.empty:
            comparable = [str(value) for value in non_null.tolist()]
            min_value = min(comparable)
            max_value = max(comparable)
    except Exception:
        min_value = None
        max_value = None

    return {
        "name": column,
        "normalized_name": _normalize_source_column(column),
        "physical_dtype": str(series.dtype),
        "semantic_type_hint": canonical_target_for_column(column),
        "nullable": bool(series.isna().any()),
        "null_count": int(series.isna().sum()),
        "distinct_count": len(distinct_values),
        "sample_values": bounded_distinct[:MAX_SAMPLE_VALUES],
        "min_value": min_value,
        "max_value": max_value,
        "detected_type": detected_type,
    }


def _load_tabular_preview(path: Path, extension: str, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    total_rows = 0
    if extension == ".csv":
        frame = pd.read_csv(path, nrows=max_preview_rows)
        total_rows = _count_delimited_rows(path)
    elif extension == ".tsv":
        frame = pd.read_csv(path, sep="\t", nrows=max_preview_rows)
        total_rows = _count_delimited_rows(path)
    else:
        frame = pd.read_excel(path, nrows=max_preview_rows)
        total_rows = _estimate_excel_rows(path, extension)
        if total_rows <= 0:
            total_rows = len(frame)
    frame = frame.dropna(how="all")
    frame.columns = [_normalize_source_column(column) for column in frame.columns]
    if total_rows <= 0:
        total_rows = len(frame)
    return frame, total_rows


def _load_json_frame(path: Path, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] | None = None
    if isinstance(payload, list) and payload and all(isinstance(item, dict) for item in payload):
        rows = payload
    elif isinstance(payload, dict):
        if payload and all(not isinstance(value, (list, dict)) for value in payload.values()):
            rows = [payload]
        else:
            nested_rows = next(
                (value for value in payload.values() if isinstance(value, list) and value and all(isinstance(item, dict) for item in value)),
                None,
            )
            if nested_rows is not None:
                rows = nested_rows
    if not rows:
        return None, 0
    total_rows = len(rows)
    return pd.DataFrame(rows[:max_preview_rows]), total_rows


def _load_text_frame(path: Path, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    lines = [line.strip() for line in path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]
    if not lines:
        return None, 0
    total_rows = len(lines)
    preview_lines = lines[:max_preview_rows]
    return pd.DataFrame([{"line_number": index + 1, "raw_text": line} for index, line in enumerate(preview_lines)]), total_rows


def _normalize_source_column(value: Any) -> str:
    raw = str(value).strip()
    alias_key = raw.lower()
    if alias_key in COLUMN_ALIASES:
        return COLUMN_ALIASES[alias_key]
    normalized = re.sub(r"[^a-z0-9]+", "_", alias_key).strip("_")
    return normalized or "column"


def _infer_detected_types(frame: pd.DataFrame) -> dict[str, str]:
    detected: dict[str, str] = {}
    for column in frame.columns:
        series = frame[column].dropna()
        if series.empty:
            detected[str(column)] = "empty"
            continue
        if pd.api.types.is_numeric_dtype(series):
            detected[str(column)] = "number"
            continue
        numeric_ratio = pd.to_numeric(series, errors="coerce").notna().mean()
        detected[str(column)] = "number" if numeric_ratio >= 0.9 else "string"
    return detected


def _build_field_mapping(source_column: str, detected_type: str) -> dict[str, str]:
    alias_match = source_column in COLUMN_ALIASES.values()
    if alias_match:
        target = source_column
        confidence = "high"
        reason = "Matched a known finance schema alias."
    elif canonical_target_for_column(source_column):
        target = canonical_target_for_column(source_column) or source_column
        confidence = "high"
        reason = "Matched a semantic column role and mapped to the canonical target."
    elif source_column == "raw_text":
        target = "raw_text"
        confidence = "medium"
        reason = "Preserved unstructured text for user confirmation."
    else:
        target = source_column
        confidence = "medium"
        reason = "Normalized from the source header and proposed as-is."
    return {
        "source": source_column,
        "target": target,
        "detected_type": detected_type,
        "confidence": confidence,
        "reason": reason,
    }


def _records_from_frame(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return frame.to_dict(orient="records")


def _count_delimited_rows(path: Path) -> int:
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            line_count = sum(1 for line in handle if line.strip())
        return max(0, line_count - 1)
    except Exception:
        return 0


def _estimate_excel_rows(path: Path, extension: str) -> int:
    try:
        if extension == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.worksheets[0] if workbook.worksheets else None
            return max(0, (sheet.max_row if sheet else 0) - 1)
        frame = pd.read_excel(path)
        return len(frame)
    except Exception:
        return 0
